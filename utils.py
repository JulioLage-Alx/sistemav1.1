import re
import csv
import os
import json
from datetime import datetime, timedelta
import decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import logging

logger = logging.getLogger(__name__)

def validar_cpf(cpf):
    """Validar CPF brasileiro"""
    if not cpf:
        return True  # CPF é opcional
    
    # Remover caracteres não numéricos
    cpf = re.sub(r'[^0-9]', '', cpf)
    
    # Verificar se tem 11 dígitos
    if len(cpf) != 11:
        return False
    
    # Verificar se não são todos iguais
    if cpf == cpf[0] * 11:
        return False
    
    # Validar primeiro dígito verificador
    soma = 0
    for i in range(9):
        soma += int(cpf[i]) * (10 - i)
    resto = soma % 11
    if resto < 2:
        digito1 = 0
    else:
        digito1 = 11 - resto
    
    if int(cpf[9]) != digito1:
        return False
    
    # Validar segundo dígito verificador
    soma = 0
    for i in range(10):
        soma += int(cpf[i]) * (11 - i)
    resto = soma % 11
    if resto < 2:
        digito2 = 0
    else:
        digito2 = 11 - resto
    
    if int(cpf[10]) != digito2:
        return False
    
    return True

def convert_decimals_to_float(data):
    if isinstance(data, list):
            return [convert_decimals_to_float(item) for item in data]
    elif isinstance(data, dict):
            return {key: convert_decimals_to_float(value) for key, value in data.items()}
    elif isinstance(data, decimal.Decimal):
            return float(data)
    else:
            return data
def formatar_moeda(valor):
    """Formatar valor monetário para exibição"""
    try:
        valor = float(valor)
        return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return "R$ 0,00"

def formatar_telefone(telefone):
    """Formatar telefone para exibição"""
    if not telefone:
        return ""
    
    # Remover caracteres não numéricos
    telefone = re.sub(r'[^0-9]', '', telefone)
    
    if len(telefone) == 11:  # Celular com DDD
        return f"({telefone[:2]}) {telefone[2:7]}-{telefone[7:]}"
    elif len(telefone) == 10:  # Fixo com DDD
        return f"({telefone[:2]}) {telefone[2:6]}-{telefone[6:]}"
    else:
        return telefone

def calcular_troco(valor_pago, valor_devido):
    """Calcular troco"""
    try:
        valor_pago = float(valor_pago)
        valor_devido = float(valor_devido)
        
        if valor_pago > valor_devido:
            return valor_pago - valor_devido
        return 0
    except:
        return 0

def validar_email(email):
    """Validar formato de email"""
    if not email:
        return True  # Email é opcional
    
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def formatar_data_brasileira(data):
    """Formatar data para padrão brasileiro"""
    if isinstance(data, str):
        return data
    
    try:
        return data.strftime('%d/%m/%Y')
    except:
        return ""

def formatar_data_hora_brasileira(data):
    """Formatar data e hora para padrão brasileiro"""
    if isinstance(data, str):
        return data
    
    try:
        return data.strftime('%d/%m/%Y %H:%M')
    except:
        return ""

def converter_data_brasileira_para_mysql(data_str):
    """Converter data do formato brasileiro (dd/mm/yyyy) para MySQL (yyyy-mm-dd)"""
    try:
        if not data_str:
            return None
        
        # Remover espaços
        data_str = data_str.strip()
        
        # Se já está no formato MySQL, retornar
        if re.match(r'^\d{4}-\d{2}-\d{2}', data_str):
            return data_str
        
        # Converter do formato brasileiro
        if '/' in data_str:
            partes = data_str.split('/')
            if len(partes) == 3:
                dia, mes, ano = partes
                return f"{ano}-{mes.zfill(2)}-{dia.zfill(2)}"
        
        return None
    except:
        return None

def sanitizar_string(texto):
    """Sanitizar string removendo caracteres especiais"""
    if not texto:
        return ""
    
    # Remover caracteres especiais mantendo apenas letras, números, espaços e acentos
    texto_limpo = re.sub(r'[^\w\s\-\.\,áéíóúâêîôûàèìòùäëïöüçãõ]', '', texto, flags=re.IGNORECASE)
    
    # Remover espaços múltiplos
    texto_limpo = re.sub(r'\s+', ' ', texto_limpo)
    
    return texto_limpo.strip()

def validar_numero_positivo(valor):
    """Validar se um valor é um número positivo"""
    try:
        numero = float(valor)
        return numero > 0
    except:
        return False

def gerar_nome_arquivo_backup():
    """Gerar nome de arquivo para backup"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"backup_acougue_{timestamp}.sql"

def criar_diretorio_se_nao_existe(caminho):
    """Criar diretório se não existir"""
    try:
        if not os.path.exists(caminho):
            os.makedirs(caminho)
        return True
    except Exception as e:
        logger.error(f"Erro ao criar diretório {caminho}: {e}")
        return False

def exportar_para_csv(dados, nome_arquivo, colunas=None):
    """Exportar dados para arquivo CSV"""
    try:
        # Criar diretório exports se não existir
        criar_diretorio_se_nao_existe('exports')
        
        caminho_arquivo = os.path.join('exports', nome_arquivo)
        
        if not dados:
            return False, "Nenhum dado para exportar"
        
        # Se não especificar colunas, usar as chaves do primeiro item
        if not colunas:
            colunas = list(dados[0].keys())
        
        with open(caminho_arquivo, 'w', newline='', encoding='utf-8') as arquivo_csv:
            writer = csv.DictWriter(arquivo_csv, fieldnames=colunas, delimiter=';')
            writer.writeheader()
            
            for linha in dados:
                # Formatar valores monetários
                linha_formatada = {}
                for chave, valor in linha.items():
                    if isinstance(valor, (int, float)) and 'valor' in chave.lower():
                        linha_formatada[chave] = formatar_moeda(valor)
                    else:
                        linha_formatada[chave] = valor
                
                writer.writerow(linha_formatada)
        
        return True, caminho_arquivo
        
    except Exception as e:
        logger.error(f"Erro ao exportar CSV: {e}")
        return False, f"Erro ao exportar: {str(e)}"

def exportar_para_excel(dados, nome_arquivo, nome_planilha="Dados", colunas=None):
    """Exportar dados para arquivo Excel"""
    try:
        # Criar diretório exports se não existir
        criar_diretorio_se_nao_existe('exports')
        
        caminho_arquivo = os.path.join('exports', nome_arquivo)
        
        if not dados:
            return False, "Nenhum dado para exportar"
        
        # Se não especificar colunas, usar as chaves do primeiro item
        if not colunas:
            colunas = list(dados[0].keys())
        
        # Criar workbook e worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = nome_planilha
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escrever cabeçalho
        for col_num, coluna in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_num, value=coluna)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Escrever dados
        for row_num, linha in enumerate(dados, 2):
            for col_num, coluna in enumerate(colunas, 1):
                valor = linha.get(coluna, '')
                
                # Formatar valores monetários
                if isinstance(valor, (int, float)) and 'valor' in coluna.lower():
                    valor = float(valor)
                
                ws.cell(row=row_num, column=col_num, value=valor)
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar arquivo
        wb.save(caminho_arquivo)
        
        return True, caminho_arquivo
        
    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {e}")
        return False, f"Erro ao exportar: {str(e)}"

def validar_arquivo_upload(arquivo, extensoes_permitidas, tamanho_max_mb=10):
    """Validar arquivo de upload"""
    if not arquivo or arquivo.filename == '':
        return False, "Nenhum arquivo selecionado"
    
    # Verificar extensão
    extensao = arquivo.filename.rsplit('.', 1)[1].lower() if '.' in arquivo.filename else ''
    if extensao not in extensoes_permitidas:
        return False, f"Extensão não permitida. Use: {', '.join(extensoes_permitidas)}"
    
    # Verificar tamanho (se possível)
    try:
        arquivo.seek(0, 2)  # Ir para o final do arquivo
        tamanho = arquivo.tell()
        arquivo.seek(0)  # Voltar para o início
        
        tamanho_mb = tamanho / (1024 * 1024)
        if tamanho_mb > tamanho_max_mb:
            return False, f"Arquivo muito grande. Máximo: {tamanho_max_mb}MB"
    except:
        pass  # Se não conseguir verificar o tamanho, prosseguir
    
    return True, "Arquivo válido"

def gerar_relatorio_json(dados, nome_arquivo):
    """Gerar relatório em formato JSON"""
    try:
        # Criar diretório exports se não existir
        criar_diretorio_se_nao_existe('exports')
        
        caminho_arquivo = os.path.join('exports', nome_arquivo)
        
        # Converter datetime para string nos dados
        dados_json = json.loads(json.dumps(dados, default=str, ensure_ascii=False, indent=2))
        
        with open(caminho_arquivo, 'w', encoding='utf-8') as arquivo_json:
            json.dump(dados_json, arquivo_json, ensure_ascii=False, indent=2)
        
        return True, caminho_arquivo
        
    except Exception as e:
        logger.error(f"Erro ao gerar JSON: {e}")
        return False, f"Erro ao gerar relatório: {str(e)}"

def calcular_idade(data_nascimento):
    """Calcular idade a partir da data de nascimento"""
    try:
        if isinstance(data_nascimento, str):
            data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d')
        
        hoje = datetime.now()
        idade = hoje.year - data_nascimento.year
        
        # Verificar se já fez aniversário este ano
        if hoje.month < data_nascimento.month or (hoje.month == data_nascimento.month and hoje.day < data_nascimento.day):
            idade -= 1
        
        return idade
    except:
        return None

def formatar_numero_quantidade(quantidade):
    """Formatar número de quantidade (peso/quantidade)"""
    try:
        quantidade = float(quantidade)
        
        # Se for um número inteiro, mostrar sem casas decimais
        if quantidade == int(quantidade):
            return str(int(quantidade))
        
        # Senão, mostrar com até 3 casas decimais
        return f"{quantidade:.3f}".rstrip('0').rstrip('.')
    except:
        return "0"

def validar_cnpj(cnpj):
    """Validar CNPJ brasileiro"""
    if not cnpj:
        return True  # CNPJ é opcional
    
    # Remover caracteres não numéricos
    cnpj = re.sub(r'[^0-9]', '', cnpj)
    
    # Verificar se tem 14 dígitos
    if len(cnpj) != 14:
        return False
    
    # Verificar se não são todos iguais
    if cnpj == cnpj[0] * 14:
        return False
    
    # Validar dígitos verificadores
    def calcular_digito(cnpj_base, pesos):
        soma = sum(int(cnpj_base[i]) * pesos[i] for i in range(len(pesos)))
        resto = soma % 11
        return 0 if resto < 2 else 11 - resto
    
    # Primeiro dígito
    pesos1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    digito1 = calcular_digito(cnpj[:12], pesos1)
    
    if int(cnpj[12]) != digito1:
        return False
    
    # Segundo dígito
    pesos2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    digito2 = calcular_digito(cnpj[:13], pesos2)
    
    if int(cnpj[13]) != digito2:
        return False
    
    return True

def obter_extensao_arquivo(nome_arquivo):
    """Obter extensão de um arquivo"""
    return nome_arquivo.rsplit('.', 1)[1].lower() if '.' in nome_arquivo else ''

def gerar_hash_simples(texto):
    """Gerar hash simples para senhas (para ambiente local apenas)"""
    import hashlib
    return hashlib.md5(texto.encode()).hexdigest()

def limpar_arquivo_temporario(caminho_arquivo, idade_minutos=60):
    """Limpar arquivos temporários antigos"""
    try:
        if os.path.exists(caminho_arquivo):
            tempo_arquivo = os.path.getmtime(caminho_arquivo)
            tempo_atual = datetime.now().timestamp()
            
            diferenca_minutos = (tempo_atual - tempo_arquivo) / 60
            
            if diferenca_minutos > idade_minutos:
                os.remove(caminho_arquivo)
                return True
        
        return False
    except Exception as e:
        logger.error(f"Erro ao limpar arquivo temporário: {e}")
        return False

def formatar_cpf_display(cpf):
    """Formatar CPF para exibição"""
    if not cpf or len(cpf) != 11:
        return cpf or ""
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"

def formatar_cnpj_display(cnpj):
    """Formatar CNPJ para exibição"""
    if not cnpj or len(cnpj) != 14:
        return cnpj or ""
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

def calcular_percentual(parte, total):
    """Calcular percentual"""
    try:
        if total == 0:
            return 0
        return (parte / total) * 100
    except:
        return 0

def obter_primeiro_nome(nome_completo):
    """Obter primeiro nome de um nome completo"""
    if not nome_completo:
        return ""
    return nome_completo.split()[0]

def validar_horario_funcionamento():
    """Verificar se está dentro do horário de funcionamento (opcional)"""
    agora = datetime.now()
    hora_atual = agora.hour
    
    # Definir horário de funcionamento (6h às 22h)
    horario_inicio = 6
    horario_fim = 22
    
    return horario_inicio <= hora_atual <= horario_fim

def gerar_numero_venda_sequencial():
    """Gerar número sequencial para venda (baseado em timestamp)"""
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    return int(timestamp)

class Logger:
    """Classe para logging customizado"""
    
    @staticmethod
    def setup_logging(nivel=logging.INFO):
        """Configurar logging"""
        logging.basicConfig(
            level=nivel,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('logs/sistema.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        # Criar diretório de logs se não existir
        criar_diretorio_se_nao_existe('logs')
    
    @staticmethod
    def log_operacao(operacao, detalhes="", usuario="sistema"):
        """Log de operação do sistema"""
        logger = logging.getLogger('operacoes')
        logger.info(f"{operacao} - {detalhes} - Usuário: {usuario}")